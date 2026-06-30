from __future__ import annotations

import argparse
import bisect
import csv
import queue
import re
import sys
import threading
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

from .document import extract_layout_pages
from .layout import extract_layout_groups, group_spans_into_lines
from .models import (
    DocumentPage,
    ExtractResult,
    ExtractionOptions,
    ProductBlock,
    ReviewCandidate,
    SerialRecord,
)
from .ocr import (
    DEFAULT_OCR_MANIFEST_URL,
    OcrEngine,
    OcrInstallError,
    OcrSupportManager,
    OcrSupportRequired,
    discover_ocr_engine,
    partition_ocr_spans,
)
from . import __version__


APP_NAME = "Serial Number Extractor"


KNOWN_SERIAL_RE = re.compile(
    r"\b(?:"
    r"LBADVA[A-Z0-9]{6,}|"
    r"FA(?=[A-Z0-9]*\d)[A-Z0-9]{8,}|"
    r"VB\d{6,}|"
    r"VK\d{6,}|"
    r"WX\d{8,}|"
    r"INC[A-Z0-9]{8,}|"
    r"FBXN[A-Z0-9]{8,}|"
    r"M9[A-Z0-9]{6,}|"
    r"S2S\d{5,}|"
    r"S4S\d{5,}|"
    r"B\d{10,}|"
    r"PLS[A-Z0-9]{3,}|"
    r"PM[A-Z0-9]{4,}|"
    r"NR[A-Z0-9]{4,}|"
    r"PHF[A-Z0-9]{4,}"
    r")\b",
    re.IGNORECASE,
)

NUMERIC_CONTEXT_RE = re.compile(r"\b\d{8,14}\b")
GENERIC_CANDIDATE_RE = re.compile(r"\b[A-Z0-9][A-Z0-9._/+:-]{4,31}\b", re.IGNORECASE)
RANGE_RE = re.compile(r"\b([A-Z]+)(\d{8,})\s*-\s*(\d{1,})\b", re.IGNORECASE)
ADTRAN_PART_RE = re.compile(r"^\d{1,5}\s+([A-Z0-9][A-Z0-9/#.+_-]{3,})$", re.IGNORECASE)
ADTRAN_QTY_RE = re.compile(r"^(\d+)\s+\d+\s+\d+\s+[\d.,]+\s+[\d.,]+$")
ADTRAN_POSITION_LINE_RE = re.compile(r"^(\d+)\s+([A-Z0-9][A-Z0-9/#.+_-]{3,})\s+(.+)$", re.IGNORECASE)
CIENA_PART_RE = re.compile(r"^\d+\s+(\d+)\s+EA\s+([A-Z0-9][A-Z0-9/#.+_-]+)$", re.IGNORECASE)
DTC_QTY_RE = re.compile(r"^\d{1,5}$")
PURE_IT_RE = re.compile(r"^(.+?)\s+\(Quantit[eé]\s+(\d+)\)\s*:\s*(.+)$", re.IGNORECASE)
RUID_ITEM_RE = re.compile(r"Item\s+code\s+([A-Z0-9-]+)", re.IGNORECASE)
SMARTOPTICS_RE = re.compile(r"^([A-Z0-9][A-Z0-9/_+.-]*?)\d{5,}\s+(\d+)\s+PCS", re.IGNORECASE)

SERIAL_MARKER_RE = re.compile(
    r"(S/N|Serial\s*(?:number|#|no)|LOT/SERIAL|SN\s*number|Serial\s*numbers\s*are)",
    re.IGNORECASE,
)
S_N_MARKER_RE = re.compile(r"\bS/N\b|\bSN\s*:", re.IGNORECASE)

ITEM_HINT_RE = re.compile(
    r"\b("
    r"\d{10}(?:-\d{2})?|"
    r"\d{3}-\d{4}-\d{3}|"
    r"XCVR-[A-Z0-9-]+|"
    r"SFP[+/A-Z0-9#()._-]*|"
    r"QSFP[+/A-Z0-9#()._-]*|"
    r"CFP[+/A-Z0-9#()._-]*|"
    r"SO-[A-Z0-9-]+|"
    r"TQD[A-Z0-9-]+|"
    r"OTR[A-Z0-9_]+"
    r")\b",
    re.IGNORECASE,
)

EXCLUDED_TOKENS = {
    "PACKING",
    "INVOICE",
    "CUSTOMER",
    "REFERENCE",
    "DESCRIPTION",
    "QUANTITY",
    "INCOTERMS",
    "SINGAPORE",
    "APEIRO",
    "NETWORKS",
    "ADTRAN",
    "ADVA",
    "SMARTOPTICS",
    "GENERAL",
    "BUSINESS",
    "PRODUCT",
    "SERIAL",
    "NUMBER",
    "DELIVERY",
    "ORDER",
    "CONTACT",
    "ADDRESS",
}

EXCLUDED_PREFIXES = (
    "SFP",
    "QSFP",
    "CFP",
    "XCVR",
    "FSP",
    "DCP",
    "TQD",
    "OTR",
    "XG",
    "GE",
    "CIE",
    "DHL",
    "GST",
    "UEN",
)

EXCLUDED_CANDIDATE_CONTEXTS = (
    "ACTIVATION ID",
    "ACCOUNT",
    "BANK",
    "BENEFICIARY",
    "BIC",
    "COMMODITY CODE",
    "COUNTRY OF ORIGIN",
    "IBAN",
    "INVOICE",
    "ORDER NO",
    "ORDER NUMBER",
    "OUR REF",
    "PACKING LIST",
    "PAYMENT",
    "PRICE",
    "ROUTING",
    "SWIFT",
    "TOTAL",
    "WIRE TRANSFER",
)


def log_noop(_: str) -> None:
    return None


def cli_log(message: str) -> None:
    try:
        print(message)
    except Exception:
        return


def write_error_log(output_xlsx: Path | None, exc: Exception) -> Path:
    base_dir = output_xlsx.parent if output_xlsx else Path.cwd()
    base_dir.mkdir(parents=True, exist_ok=True)
    log_path = base_dir / "serial_number_extractor_error.log"
    log_path.write_text(traceback.format_exc(), encoding="utf-8")
    return log_path


def import_runtime_dependencies():
    try:
        from pypdf import PdfReader  # type: ignore
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font, PatternFill  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency. This app needs pypdf and openpyxl. "
            "Use the bundled launcher or install dependencies with: "
            "python -m pip install pypdf openpyxl"
        ) from exc
    return PdfReader, Workbook, Font, PatternFill, get_column_letter


def normalize_serial(value: str) -> str:
    value = value.strip().strip(",;:.()[]{}<>")
    value = value.strip("*")
    value = re.sub(r"^(?:S/N|SN)\s*:?", "", value, flags=re.IGNORECASE)
    value = re.sub(r"^(FA\d{8,})SN$", r"\1", value, flags=re.IGNORECASE)
    return value.upper()


def is_probable_serial(token: str, allow_numeric: bool = False) -> bool:
    token = normalize_serial(token)
    if not token:
        return False
    if token in EXCLUDED_TOKENS:
        return False
    if token.startswith(EXCLUDED_PREFIXES):
        return False
    if len(token) < 5 or len(token) > 32:
        return False
    if re.fullmatch(r"\d+", token):
        return allow_numeric and 10 <= len(token) <= 14
    if "-" in token:
        return False
    return bool(re.search(r"[A-Z]", token) and re.search(r"\d", token))


def is_excluded_candidate_line(line: str) -> bool:
    upper = line.upper()
    return any(context in upper for context in EXCLUDED_CANDIDATE_CONTEXTS)


def has_usable_native_page_text(text: str, lines: list[str]) -> bool:
    if sum(character.isalnum() for character in text) >= 20:
        return True
    if KNOWN_SERIAL_RE.search(text) or RANGE_RE.search(text):
        return True
    return any(SERIAL_MARKER_RE.search(line) or S_N_MARKER_RE.search(line) for line in lines)


def infer_range_end(start_digits: str, end_suffix: str) -> int | None:
    if len(end_suffix) > len(start_digits):
        return None
    prefix = start_digits[: len(start_digits) - len(end_suffix)]
    candidate = int(prefix + end_suffix)
    start_num = int(start_digits)
    if candidate < start_num:
        if not prefix:
            return None
        incremented_prefix = str(int(prefix) + 1)
        candidate = int(incremented_prefix + end_suffix)
    return candidate if candidate >= start_num else None


def expand_serial_range(raw: str, max_items: int = 5000) -> list[str]:
    match = RANGE_RE.fullmatch(raw.strip())
    if not match:
        return []
    prefix, start_digits, end_suffix = match.groups()
    end_num = infer_range_end(start_digits, end_suffix)
    if end_num is None:
        return []
    start_num = int(start_digits)
    count = end_num - start_num + 1
    if count <= 0 or count > max_items:
        return []
    width = len(start_digits)
    return [f"{prefix.upper()}{str(value).zfill(width)}" for value in range(start_num, end_num + 1)]


def clean_text_lines(text: str) -> list[str]:
    lines = []
    for raw in text.splitlines():
        line = re.sub(r"\s+", " ", raw).strip()
        if line:
            lines.append(line)
    return lines


def line_offsets(lines: list[str]) -> list[int]:
    offsets: list[int] = []
    cursor = 0
    for line in lines:
        offsets.append(cursor)
        cursor += len(line) + 1
    return offsets


def line_index_from_offset(offsets: list[int], offset: int) -> int:
    if not offsets:
        return 0
    return max(0, min(len(offsets) - 1, bisect.bisect_right(offsets, offset) - 1))


def first_int(value: str) -> int | None:
    match = re.search(r"\d+", value)
    return int(match.group(0)) if match else None


def clean_part_value(value: str) -> str:
    value = re.sub(r"\s+", " ", value).strip()
    return value.strip(" ,:;")


def strip_serials_from_part_name(value: str) -> str:
    value = KNOWN_SERIAL_RE.sub("", value)
    value = re.sub(r"\s*,\s*", " ", value)
    return clean_part_value(value)


def is_part_number_like(value: str) -> bool:
    value = clean_part_value(value)
    if not value or len(value) > 42:
        return False
    if value.upper() in EXCLUDED_TOKENS:
        return False
    if re.search(r"\s", value):
        return False
    if re.fullmatch(r"\d{6,10}", value):
        return True
    if re.fullmatch(r"BC\d{8}", value, re.IGNORECASE):
        return True
    if is_probable_serial(value, allow_numeric=True) and not re.search(r"[/#.+_-]", value):
        return False
    if re.fullmatch(r"\d{1,5}", value):
        return False
    return bool(re.search(r"\d", value) or re.search(r"[-_/+#.]", value))


def is_valid_part_name(value: str) -> bool:
    value = clean_part_value(value)
    if not value:
        return False
    upper = value.upper()
    bad_terms = (
        "COMMODITY CODE",
        "COUNTRY OF ORIGIN",
        "PACKING LIST",
        "INVOICE",
        "PRICE PER",
        "UNIT EUR",
        "TOTAL",
        "SERIAL",
        "S/N",
        "PAGE OF",
        "CUSTOMER",
        "REFERENCE",
        "ADDRESS",
    )
    if any(term in upper for term in bad_terms):
        return False
    if is_probable_serial(value, allow_numeric=True) and not re.search(r"[/#.+_-]", value):
        return False
    if re.fullmatch(r"[\d\s.,-]+", value):
        return False
    return True


def add_product_block(
    blocks: list[ProductBlock],
    start_line: int,
    end_line: int,
    part_number: str,
    part_name: str,
    order_qty: int | None,
    source: str,
) -> None:
    part_number = clean_part_value(part_number).upper()
    part_name = clean_part_value(part_name)
    if not part_number and not part_name:
        return
    key = f"{source}:{start_line}:{part_number}:{part_name}:{order_qty or ''}"
    if any(block.block_key == key for block in blocks):
        return
    blocks.append(
        ProductBlock(
            part_number=part_number,
            part_name=part_name,
            order_qty=order_qty,
            start_line=start_line,
            end_line=end_line,
            block_key=key,
            source=source,
        )
    )


def next_matching_line(lines: list[str], start: int, pattern: re.Pattern[str]) -> int:
    for idx in range(start + 1, len(lines)):
        if pattern.search(lines[idx]):
            return idx
    return len(lines)


def find_adtran_qty(lines: list[str], start: int, end: int) -> int | None:
    for idx in range(start + 1, min(end + 1, len(lines))):
        if "THIS POSITION LINE CONTAINS" in lines[idx].upper():
            for position_idx in range(idx + 1, min(end + 1, len(lines))):
                parsed = parse_adtran_position_line(lines[position_idx])
                if parsed:
                    return parsed[0]
                if ADTRAN_QTY_RE.match(lines[position_idx]) or ADTRAN_PART_RE.match(lines[position_idx]):
                    break
        match = ADTRAN_QTY_RE.match(lines[idx])
        if match:
            return int(match.group(1))
    return None


def parse_adtran_position_line(line: str) -> tuple[int, str, str] | None:
    match = ADTRAN_POSITION_LINE_RE.match(line)
    if not match:
        return None
    qty = int(match.group(1))
    part_number = clean_part_value(match.group(2))
    part_name = strip_serials_from_part_name(match.group(3))
    if not is_part_number_like(part_number) or not is_valid_part_name(part_name):
        return None
    return qty, part_number, part_name


def finalize_product_blocks(blocks: list[ProductBlock], line_count: int) -> list[ProductBlock]:
    blocks.sort(key=lambda block: (block.start_line, block.end_line))
    final: list[ProductBlock] = []
    for idx, block in enumerate(blocks):
        next_start = blocks[idx + 1].start_line if idx + 1 < len(blocks) else line_count
        block.end_line = min(block.end_line, next_start - 1, line_count - 1)
        if block.end_line < block.start_line:
            block.end_line = block.start_line
        final.append(block)
    return final


def detect_pure_it_blocks(lines: list[str], blocks: list[ProductBlock]) -> None:
    for idx, line in enumerate(lines):
        match = PURE_IT_RE.match(line)
        if match:
            part_number = match.group(1)
            qty = int(match.group(2))
            part_name = match.group(3)
            end = next_matching_line(lines, idx, PURE_IT_RE) - 1
            add_product_block(blocks, idx, end, part_number, part_name, qty, "pure_it")


def detect_smartoptics_blocks(lines: list[str], blocks: list[ProductBlock]) -> None:
    for idx, line in enumerate(lines):
        match = SMARTOPTICS_RE.match(line)
        if match and idx + 1 < len(lines):
            part_number = match.group(1)
            qty = int(match.group(2))
            if idx + 2 < len(lines) and is_part_number_like(lines[idx + 2]):
                part_number = lines[idx + 2]
            part_name = strip_serials_from_part_name(lines[idx + 1])
            if is_valid_part_name(part_name):
                end = next_matching_line(lines, idx, SMARTOPTICS_RE) - 1
                add_product_block(blocks, idx, end, part_number, part_name, qty, "smartoptics")


def detect_dtc_blocks(lines: list[str], blocks: list[ProductBlock]) -> None:
    for idx in range(0, max(0, len(lines) - 4)):
        if (
            is_valid_part_name(lines[idx])
            and lines[idx + 1].upper().startswith("ADVA")
            and is_part_number_like(lines[idx + 2])
            and DTC_QTY_RE.match(lines[idx + 3])
            and "Serial numbers are" in lines[idx + 4]
        ):
            end = next((j - 1 for j in range(idx + 5, len(lines) - 4) if lines[j + 4].startswith("Serial numbers are:")), len(lines) - 1)
            add_product_block(blocks, idx, end, lines[idx + 2], lines[idx], int(lines[idx + 3]), "dtc")


def detect_ruid_blocks(lines: list[str], blocks: list[ProductBlock]) -> None:
    for idx in range(0, max(0, len(lines) - 4)):
        item_match = RUID_ITEM_RE.search(lines[idx + 2])
        if item_match and idx + 4 < len(lines) and RANGE_RE.search(lines[idx + 4] + " " + (lines[idx + 5] if idx + 5 < len(lines) else "")):
            qty = first_int(lines[idx + 4])
            part_name = lines[idx + 1] if is_valid_part_name(lines[idx + 1]) else lines[idx]
            end = idx + 5
            add_product_block(blocks, idx, end, item_match.group(1), part_name, qty, "ruid")


def detect_ciena_blocks(lines: list[str], blocks: list[ProductBlock]) -> None:
    for idx, line in enumerate(lines):
        match = CIENA_PART_RE.match(line)
        if match:
            qty = int(match.group(1))
            part_number = match.group(2)
            part_name_parts: list[str] = []
            for name_idx in range(idx + 1, min(idx + 4, len(lines))):
                if CIENA_PART_RE.match(lines[name_idx]):
                    break
                candidate = strip_serials_from_part_name(lines[name_idx])
                if is_valid_part_name(candidate):
                    part_name_parts.append(candidate)
                if KNOWN_SERIAL_RE.search(lines[name_idx]):
                    break
            part_name = " ".join(part_name_parts) or part_number
            end = next_matching_line(lines, idx, CIENA_PART_RE) - 1
            add_product_block(blocks, idx, end, part_number, part_name, qty, "ciena")


def detect_adtran_blocks(lines: list[str], blocks: list[ProductBlock]) -> None:
    for idx, line in enumerate(lines):
        match = ADTRAN_PART_RE.match(line)
        if not match or idx + 1 >= len(lines):
            continue
        part_number = match.group(1)
        part_name = lines[idx + 1]
        if not is_part_number_like(part_number) or not is_valid_part_name(part_name):
            continue
        end = next_matching_line(lines, idx, ADTRAN_PART_RE) - 1
        qty = find_adtran_qty(lines, idx, end)
        add_product_block(blocks, idx, end, part_number, part_name, qty, "adtran")


def detect_eci_best_effort_blocks(lines: list[str], blocks: list[ProductBlock]) -> None:
    if any("LOT/SERIAL" in line.upper() for line in lines):
        eci_starts = [
            idx
            for idx, line in enumerate(lines)
            if re.fullmatch(r"\d{6,10}", line) and idx + 1 < len(lines) and is_valid_part_name(lines[idx + 1])
        ]
        for position, idx in enumerate(eci_starts):
            next_start = eci_starts[position + 1] if position + 1 < len(eci_starts) else len(lines)
            add_product_block(blocks, idx, next_start - 1, lines[idx], lines[idx + 1], None, "eci_best_effort")


@dataclass(frozen=True)
class ProductBlockStrategy:
    name: str
    priority: int
    detect: Callable[[list[str], list[ProductBlock]], None]


PRODUCT_BLOCK_STRATEGIES = tuple(
    sorted(
        (
            ProductBlockStrategy("pure_it", 100, detect_pure_it_blocks),
            ProductBlockStrategy("smartoptics", 100, detect_smartoptics_blocks),
            ProductBlockStrategy("dtc", 100, detect_dtc_blocks),
            ProductBlockStrategy("ruid", 100, detect_ruid_blocks),
            ProductBlockStrategy("ciena", 100, detect_ciena_blocks),
            ProductBlockStrategy("adtran", 110, detect_adtran_blocks),
            ProductBlockStrategy("eci_best_effort", 20, detect_eci_best_effort_blocks),
        ),
        key=lambda strategy: strategy.priority,
        reverse=True,
    )
)


def build_product_blocks(lines: list[str]) -> list[ProductBlock]:
    blocks: list[ProductBlock] = []
    for strategy in PRODUCT_BLOCK_STRATEGIES:
        strategy.detect(lines, blocks)
    return finalize_product_blocks(blocks, len(lines))


def find_product_block(blocks: list[ProductBlock], line_index: int) -> ProductBlock | None:
    candidates = [block for block in blocks if block.start_line <= line_index <= block.end_line]
    if candidates:
        return max(candidates, key=lambda block: block.start_line)
    previous = [block for block in blocks if block.start_line <= line_index]
    if previous:
        block = max(previous, key=lambda item: item.start_line)
        if line_index - block.end_line <= 3:
            return block
    return None


def find_usi_code_lines(lines: list[str]) -> set[int]:
    usi_lines: set[int] = set()
    in_usi_block = False
    for idx, line in enumerate(lines):
        upper = line.upper()
        if "USI CODE" in upper:
            in_usi_block = True
            usi_lines.add(idx)
            continue
        if not in_usi_block:
            continue
        if (
            ADTRAN_QTY_RE.match(line)
            or ADTRAN_PART_RE.match(line)
            or ADTRAN_POSITION_LINE_RE.match(line)
            or SERIAL_MARKER_RE.search(line)
            or S_N_MARKER_RE.search(line)
        ):
            in_usi_block = False
            continue
        if "INVOICE DETAILS PAGE" in upper or "THIS POSITION LINE CONTAINS" in upper:
            in_usi_block = False
            continue
        usi_lines.add(idx)
    return usi_lines


def find_item_hint(lines: list[str], line_index: int) -> str:
    for idx in range(line_index, max(-1, line_index - 10), -1):
        line = lines[idx]
        matches = [m.group(1) for m in ITEM_HINT_RE.finditer(line)]
        if matches:
            return matches[-1]
    return ""


def duplicate_serial_key(
    source_file: str,
    serial: str,
    block_key: str = "",
) -> tuple[str, str]:
    serial = normalize_serial(serial)
    scope = f"{source_file}:{block_key or 'document'}"
    if serial.startswith("LBADVA"):
        return (scope, serial.removeprefix("LBADVA"))
    if serial.startswith("FA"):
        return (scope, serial.removeprefix("FA"))
    return (scope, serial)


def add_record(
    records: list[SerialRecord],
    seen: set[tuple[str, str]],
    source_file: str,
    page: int,
    serial: str,
    item_hint: str,
    method: str,
    confidence: str,
    product_block: ProductBlock | None,
) -> None:
    serial = normalize_serial(serial)
    block_key = product_block.block_key if product_block else ""
    key = duplicate_serial_key(source_file, serial, block_key)
    if not serial or key in seen:
        return
    seen.add(key)
    part_number = product_block.part_number if product_block else item_hint
    part_name = product_block.part_name if product_block else ""
    order_qty = product_block.order_qty if product_block else None
    records.append(
        SerialRecord(
            source_file=source_file,
            page=page,
            part_number=part_number,
            part_name=part_name,
            order_qty=order_qty,
            serial_number=serial,
            serial_count=None,
            qty_check="UNVERIFIED",
            item_hint=item_hint,
            method=method,
            confidence=confidence,
            block_key=block_key,
            block_source=product_block.source if product_block else "",
            strategy=product_block.source if product_block else "generic_marker",
            candidate_score=0.95 if confidence == "high" else 0.80,
        )
    )


def extract_document_records(
    source_file: str,
    lines: list[str],
    line_pages: list[int],
    seen: set[tuple[str, str]],
    restricted_pages: set[int] | None = None,
) -> list[SerialRecord]:
    records: list[SerialRecord] = []
    full_text = "\n".join(lines)
    offsets = line_offsets(lines)
    product_blocks = build_product_blocks(lines)
    restricted_pages = restricted_pages or set()
    usi_code_lines = find_usi_code_lines(lines)
    marker_lines = [
        index
        for index, line in enumerate(lines)
        if SERIAL_MARKER_RE.search(line) or S_N_MARKER_RE.search(line)
    ]
    restricted_context_indexes: set[int] = set()
    for marker_index in marker_lines:
        next_block_start = min(
            (
                block.start_line
                for block in product_blocks
                if block.start_line > marker_index
            ),
            default=len(lines),
        )
        marker_end = min(len(lines), marker_index + 18, next_block_start)
        for index in range(marker_index, marker_end):
            if index > marker_index and is_excluded_candidate_line(lines[index]):
                break
            restricted_context_indexes.add(index)

    def block_has_primary_marker(line_index: int) -> bool:
        block = find_product_block(product_blocks, line_index)
        if block is None:
            return any(abs(marker_index - line_index) <= 24 for marker_index in marker_lines)
        return any(block.start_line <= marker_index <= block.end_line for marker_index in marker_lines)

    def allowed_on_restricted_page(line_index: int) -> bool:
        page = line_pages[line_index] if line_index < len(line_pages) else 1
        if page not in restricted_pages:
            return True
        if line_index in restricted_context_indexes:
            return True
        product_block = find_product_block(product_blocks, line_index)
        if product_block is not None and product_block.source != "eci_best_effort":
            return True
        return line_index in usi_code_lines and not block_has_primary_marker(line_index)

    for match in RANGE_RE.finditer(full_text):
        raw_range = match.group(0)
        expanded = expand_serial_range(raw_range)
        if not expanded:
            continue
        line_index = line_index_from_offset(offsets, match.start())
        if not allowed_on_restricted_page(line_index):
            continue
        item_hint = find_item_hint(lines, line_index)
        product_block = find_product_block(product_blocks, line_index)
        page = line_pages[line_index] if line_index < len(line_pages) else 1
        for serial in expanded:
            add_record(records, seen, source_file, page, serial, item_hint, "range_expanded", "high", product_block)

    for line_index, line in enumerate(lines):
        if not allowed_on_restricted_page(line_index):
            continue
        if line_index in usi_code_lines and block_has_primary_marker(line_index):
            continue
        item_hint = find_item_hint(lines, line_index)
        product_block = find_product_block(product_blocks, line_index)
        page = line_pages[line_index] if line_index < len(line_pages) else 1
        for match in KNOWN_SERIAL_RE.finditer(line):
            method = "usi_code" if line_index in usi_code_lines else "known_pattern"
            add_record(records, seen, source_file, page, match.group(0), item_hint, method, "high", product_block)

    for line_index in sorted(restricted_context_indexes):
        if line_index in usi_code_lines and block_has_primary_marker(line_index):
            continue
        line = lines[line_index]
        if is_excluded_candidate_line(line):
            continue
        product_block = find_product_block(product_blocks, line_index)
        page = line_pages[line_index] if line_index < len(line_pages) else 1
        for match in GENERIC_CANDIDATE_RE.finditer(line):
            candidate = normalize_serial(match.group(0))
            if not is_probable_serial(candidate):
                continue
            if product_block is not None:
                if candidate == product_block.part_number.upper():
                    continue
                if candidate in product_block.part_name.upper().split():
                    continue
            add_record(
                records,
                seen,
                source_file,
                page,
                candidate,
                find_item_hint(lines, line_index),
                "marker_alphanumeric",
                "medium",
                product_block,
            )

    allow_numeric_context = bool(re.search(r"LOT/SERIAL|\bECI\b", full_text, re.IGNORECASE))
    for idx in sorted(restricted_context_indexes):
        if idx in usi_code_lines:
            continue
        line = lines[idx]
        if is_excluded_candidate_line(line):
            continue
        if allow_numeric_context:
            product_block = find_product_block(product_blocks, idx)
            page = line_pages[idx] if idx < len(line_pages) else 1
            for match in NUMERIC_CONTEXT_RE.finditer(line):
                token = match.group(0)
                if is_probable_serial(token, allow_numeric=True):
                    add_record(records, seen, source_file, page, token, "", "numeric_serial_context", "medium", product_block)

    return records


def _layout_records(
    source_file: str,
    layout_groups,
    ocr_pages: set[int],
) -> list[SerialRecord]:
    records: list[SerialRecord] = []
    seen: set[tuple[str, str]] = set()
    for group_index, group in enumerate(layout_groups):
        block_key = (
            f"generic_layout:{group.page}:{group_index}:"
            f"{group.part_number}:{group.part_name}:{group.order_qty}"
        )
        backend = "paddleocr" if group.page in ocr_pages else "native_layout"
        for serial in group.serials:
            normalized = normalize_serial(serial)
            key = duplicate_serial_key(source_file, normalized, block_key)
            if not normalized or key in seen:
                continue
            seen.add(key)
            records.append(
                SerialRecord(
                    source_file=source_file,
                    page=group.page,
                    part_number=group.part_number,
                    part_name=group.part_name,
                    order_qty=group.order_qty,
                    serial_number=normalized,
                    serial_count=None,
                    qty_check="UNVERIFIED",
                    item_hint=group.part_number,
                    method="layout_table",
                    confidence="high" if group.confidence >= 0.90 else "medium",
                    block_key=block_key,
                    block_source="generic_layout",
                    backend=backend,
                    strategy="generic_layout",
                    candidate_score=group.confidence,
                    ocr_confidence=group.confidence if backend == "paddleocr" else None,
                    bbox=group.bbox,
                )
            )
    return records


def _merge_layout_records(
    legacy_records: list[SerialRecord],
    layout_records: list[SerialRecord],
) -> list[SerialRecord]:
    queues: dict[str, list[SerialRecord]] = {}
    for record in layout_records:
        queues.setdefault(record.serial_number, []).append(record)
    layout_pages = {record.page for record in layout_records}

    merged: list[SerialRecord] = []
    for record in legacy_records:
        if record.block_source == "eci_best_effort" and record.page in layout_pages:
            continue
        matches = queues.get(record.serial_number, [])
        if matches:
            merged.append(matches.pop(0))
        else:
            merged.append(record)
    for matches in queues.values():
        merged.extend(matches)
    return merged


def _ocr_document_pages(spans, page_numbers: set[int]) -> list[DocumentPage]:
    pages: list[DocumentPage] = []
    for page_number in sorted(page_numbers):
        page_spans = [span for span in spans if span.page == page_number]
        line_groups = group_spans_into_lines(page_spans)
        lines = [" ".join(span.text for span in line).strip() for line in line_groups]
        lines = [line for line in lines if line]
        pages.append(
            DocumentPage(
                page_number=page_number,
                text="\n".join(lines),
                lines=lines,
                spans=page_spans,
                backend="paddleocr",
                scanned=True,
            )
        )
    return pages


def _attach_review_context(
    review_candidates: list[ReviewCandidate],
    raw_ocr_spans,
    page_numbers: set[int],
) -> None:
    if not review_candidates:
        return
    groups = extract_layout_groups(_ocr_document_pages(raw_ocr_spans, page_numbers))
    for candidate in review_candidates:
        for group in groups:
            if group.page != candidate.page:
                continue
            if candidate.normalized_value not in {
                normalize_serial(serial) for serial in group.serials
            }:
                continue
            candidate.part_number = group.part_number
            candidate.part_name = group.part_name
            candidate.order_qty = group.order_qty
            candidate.strategy = group.strategy
            break


def extract_pdf(
    pdf_path: Path,
    log: Callable[[str], None] = log_noop,
    options: ExtractionOptions | None = None,
    ocr_engine: OcrEngine | None = None,
) -> ExtractResult:
    PdfReader, _, _, _, _ = import_runtime_dependencies()
    options = options or ExtractionOptions()
    records: list[SerialRecord] = []
    warnings: list[str] = []
    seen: set[tuple[str, str]] = set()

    try:
        reader = PdfReader(str(pdf_path))
    except Exception as exc:
        return ExtractResult([], [f"{pdf_path.name}: failed to open PDF: {exc}"])

    if getattr(reader, "is_encrypted", False):
        try:
            decrypt_result = reader.decrypt("")
        except Exception as exc:
            return ExtractResult([], [f"{pdf_path.name}: encrypted PDF is not readable: {exc}"])
        if not decrypt_result:
            return ExtractResult([], [f"{pdf_path.name}: encrypted PDF requires a password."])

    try:
        page_count = len(reader.pages)
    except Exception as exc:
        return ExtractResult([], [f"{pdf_path.name}: failed to read PDF pages: {exc}"])

    native_lines_by_page: dict[int, list[str]] = {}
    native_text_by_page: dict[int, str] = {}
    for page_index in range(1, page_count + 1):
        try:
            page = reader.pages[page_index - 1]
            text = page.extract_text() or ""
        except Exception as exc:
            warnings.append(f"{pdf_path.name} page {page_index}: failed to extract text: {exc}")
            native_lines_by_page[page_index] = []
            native_text_by_page[page_index] = ""
            continue
        native_text_by_page[page_index] = text
        native_lines_by_page[page_index] = clean_text_lines(text)

    if options.ocr_mode == "force":
        pages_to_ocr = set(native_lines_by_page)
    elif options.ocr_mode == "auto":
        pages_to_ocr = {
            page_number
            for page_number, text in native_text_by_page.items()
            if not has_usable_native_page_text(text, native_lines_by_page.get(page_number, []))
        }
    else:
        pages_to_ocr = set()

    review_candidates = []
    accepted_ocr_spans = []
    ocr_replaced_pages: set[int] = set()
    if pages_to_ocr and options.ocr_mode != "off":
        engine = ocr_engine or discover_ocr_engine()
        raw_ocr_spans = engine.recognize_pdf(pdf_path, sorted(pages_to_ocr))
        accepted_ocr_spans, review_candidates = partition_ocr_spans(
            source_file=pdf_path.name,
            spans=raw_ocr_spans,
            min_confidence=options.ocr_min_confidence,
        )
        _attach_review_context(review_candidates, raw_ocr_spans, pages_to_ocr)
        for page in _ocr_document_pages(accepted_ocr_spans, pages_to_ocr):
            native_text = native_text_by_page.get(page.page_number, "")
            native_lines = native_lines_by_page.get(page.page_number, [])
            if options.ocr_mode != "force" or not has_usable_native_page_text(native_text, native_lines):
                native_lines_by_page[page.page_number] = page.lines
                native_text_by_page[page.page_number] = page.text
                ocr_replaced_pages.add(page.page_number)
    elif options.ocr_mode == "off":
        scanned_pages = [
            page_number
            for page_number, text in native_text_by_page.items()
            if sum(character.isalnum() for character in text) < 20
        ]
        if scanned_pages:
            warnings.append(
                f"{pdf_path.name}: pages {', '.join(map(str, scanned_pages))} appear scanned; OCR is disabled."
            )

    lines: list[str] = []
    line_pages: list[int] = []
    for page_number in sorted(native_lines_by_page):
        page_lines = native_lines_by_page[page_number]
        lines.extend(page_lines)
        line_pages.extend([page_number] * len(page_lines))

    if lines:
        records.extend(
            extract_document_records(
                pdf_path.name,
                lines,
                line_pages,
                seen,
                restricted_pages=ocr_replaced_pages,
            )
        )
        for record in records:
            if record.page in ocr_replaced_pages:
                record.backend = "paddleocr"
                record.strategy = record.block_source or "marker"
                matching_confidences = [
                    span.confidence
                    for span in accepted_ocr_spans
                    if span.page == record.page and normalize_serial(span.text) == record.serial_number
                ]
                if matching_confidences:
                    record.ocr_confidence = max(matching_confidences)
                    record.candidate_score = record.ocr_confidence

    layout_used = False
    needs_layout = not records or not any(
        record.part_number and record.part_name and record.order_qty is not None
        for record in records
    )
    if needs_layout:
        layout_record_ocr_pages: set[int] = set()
        try:
            layout_pages = extract_layout_pages(pdf_path)
            if pages_to_ocr:
                record_pages = {record.page for record in records}
                ocr_layout_pages = {
                    page_number
                    for page_number in pages_to_ocr
                    if page_number in ocr_replaced_pages or page_number not in record_pages
                }
                ocr_page_map = {
                    page.page_number: page
                    for page in _ocr_document_pages(accepted_ocr_spans, ocr_layout_pages)
                }
                layout_record_ocr_pages = set(ocr_page_map)
                layout_pages = [
                    ocr_page_map.get(page.page_number, page)
                    for page in layout_pages
                    if page.page_number not in ocr_layout_pages or page.page_number in ocr_page_map
                ]
                for page_number, page in ocr_page_map.items():
                    if not any(existing.page_number == page_number for existing in layout_pages):
                        layout_pages.append(page)
            layout_groups = extract_layout_groups(sorted(layout_pages, key=lambda page: page.page_number))
        except Exception as exc:
            warnings.append(f"{pdf_path.name}: layout extraction failed: {exc}")
            layout_groups = []
        if layout_groups and all(len(group.serials) == group.order_qty for group in layout_groups):
            layout_records = _layout_records(pdf_path.name, layout_groups, layout_record_ocr_pages)
            if layout_records:
                records = _merge_layout_records(records, layout_records)
                layout_used = True

    if not records:
        warnings.append(f"{pdf_path.name}: no serial numbers found. This may be scanned/image-only or unsupported layout.")
    log(f"{pdf_path.name}: {len(records)} serial numbers")
    return ExtractResult(
        records,
        warnings,
        review_candidates=review_candidates,
        diagnostics={
            "layout_used": layout_used,
            "native_page_count": page_count,
            "ocr_page_count": len(pages_to_ocr),
            "ocr_replaced_page_count": len(ocr_replaced_pages),
            "review_candidate_count": len(review_candidates),
            "record_backend_counts": {
                backend: sum(1 for record in records if record.backend == backend)
                for backend in sorted({record.backend for record in records})
            },
            "record_strategy_counts": {
                strategy: sum(1 for record in records if record.strategy == strategy)
                for strategy in sorted({record.strategy for record in records})
            },
        },
    )


def collect_pdf_paths(inputs: Iterable[Path]) -> list[Path]:
    paths: list[Path] = []
    for input_path in inputs:
        if input_path.is_dir():
            paths.extend(sorted(input_path.glob("*.pdf")))
            paths.extend(sorted(input_path.glob("*.PDF")))
        elif input_path.is_file() and input_path.suffix.lower() == ".pdf":
            paths.append(input_path)
    unique: list[Path] = []
    seen = set()
    for path in paths:
        resolved = str(path.resolve()).lower()
        if resolved not in seen:
            seen.add(resolved)
            unique.append(path)
    return unique


def autosize_sheet(sheet, get_column_letter):
    for column_cells in sheet.columns:
        max_len = 0
        column = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max_len + 2, 12), 60)


def validation_group_key(record: SerialRecord) -> tuple[str, str, str, str, int | None]:
    block_key = record.block_key or f"{record.source_file}:{record.part_number}:{record.part_name}:{record.order_qty or ''}"
    return (record.source_file, block_key, record.part_number, record.part_name, record.order_qty)


def apply_quantity_validation(records: list[SerialRecord], warnings: list[str]) -> None:
    groups: dict[tuple[str, str, str, str, int | None], list[SerialRecord]] = {}
    for record in records:
        groups.setdefault(validation_group_key(record), []).append(record)

    for (source_file, _block_key, part_number, part_name, order_qty), group_records in groups.items():
        serial_count = len(group_records)
        if order_qty is None:
            status = "UNVERIFIED"
        elif serial_count == order_qty:
            status = "OK"
        else:
            status = "MISMATCH"
            label = " / ".join(value for value in [part_number, part_name] if value) or "Unknown part"
            warnings.append(f"{source_file}: {label}: order qty {order_qty}, serial count {serial_count}")
        for record in group_records:
            record.serial_count = serial_count
            record.qty_check = status


def part_summary_rows(records: list[SerialRecord]) -> list[list[object]]:
    rows: list[list[object]] = []
    seen: set[tuple[str, str, str, str, int | None]] = set()
    for record in records:
        key = validation_group_key(record)
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            [
                record.source_file,
                record.part_number,
                record.part_name,
                "" if record.order_qty is None else record.order_qty,
                "" if record.serial_count is None else record.serial_count,
                record.qty_check,
                record.block_source,
            ]
        )
    return sorted(rows, key=lambda row: (str(row[0]), str(row[1]), str(row[2]), str(row[3])))


def _bbox_text(bbox) -> str:
    if bbox is None:
        return ""
    return ",".join(f"{value:.2f}" for value in bbox)


def write_outputs(
    records: list[SerialRecord],
    warnings: list[str],
    output_xlsx: Path,
    review_candidates: list[ReviewCandidate] | None = None,
    diagnostics: dict[str, object] | None = None,
) -> tuple[Path, Path]:
    _, Workbook, Font, PatternFill, get_column_letter = import_runtime_dependencies()
    review_candidates = review_candidates or []
    diagnostics = diagnostics or {}
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_copy = wb.active
    ws_copy.title = "SAP_Copy"
    output_headers = ["Part Number", "Part Name", "Serial Number", "Order Qty", "Serial Count", "Qty Check"]
    ws_copy.append(output_headers)
    for record in records:
        ws_copy.append(
            [
                record.part_number,
                record.part_name,
                record.serial_number,
                "" if record.order_qty is None else record.order_qty,
                "" if record.serial_count is None else record.serial_count,
                record.qty_check,
            ]
        )

    ws_detail = wb.create_sheet("Details")
    ws_detail.append(
        [
            "Source File",
            "Page",
            "Part Number",
            "Part Name",
            "Order Qty",
            "Serial Count",
            "Qty Check",
            "Serial Number",
            "Item / Part Hint",
            "Method",
            "Confidence",
            "Block Source",
            "Backend",
            "Strategy",
            "Candidate Score",
            "OCR Confidence",
            "Bounding Box",
        ]
    )
    for record in records:
        ws_detail.append(
            [
                record.source_file,
                record.page,
                record.part_number,
                record.part_name,
                "" if record.order_qty is None else record.order_qty,
                "" if record.serial_count is None else record.serial_count,
                record.qty_check,
                record.serial_number,
                record.item_hint,
                record.method,
                record.confidence,
                record.block_source,
                record.backend,
                record.strategy,
                record.candidate_score,
                "" if record.ocr_confidence is None else record.ocr_confidence,
                _bbox_text(record.bbox),
            ]
        )

    ws_review = wb.create_sheet("Review")
    ws_review.append(
        [
            "Source File",
            "Page",
            "Candidate Type",
            "Raw Text",
            "Normalized Value",
            "Confidence",
            "Reason",
            "Part Number",
            "Part Name",
            "Order Qty",
            "Bounding Box",
            "Backend",
            "Strategy",
        ]
    )
    for candidate in review_candidates:
        ws_review.append(
            [
                candidate.source_file,
                candidate.page,
                candidate.candidate_type,
                candidate.raw_text,
                candidate.normalized_value,
                candidate.confidence,
                candidate.reason,
                candidate.part_number,
                candidate.part_name,
                "" if candidate.order_qty is None else candidate.order_qty,
                _bbox_text(candidate.bbox),
                candidate.backend,
                candidate.strategy,
            ]
        )

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["Total Serial Numbers", len(records)])
    ws_summary.append(["Source Files", len({record.source_file for record in records})])
    ws_summary.append(["Native Records", sum(record.backend != "paddleocr" for record in records)])
    ws_summary.append(["OCR Records", sum(record.backend == "paddleocr" for record in records)])
    ws_summary.append(["Review Candidates", len(review_candidates)])
    ws_summary.append(["OCR Pages", diagnostics.get("ocr_page_count", 0)])
    ws_summary.append(["OCR Support Installed", diagnostics.get("ocr_installed", False)])
    ws_summary.append(["OCR Support Version", diagnostics.get("ocr_version", "")])
    ws_summary.append([])
    ws_summary.append(["Source File", "Serial Count"])
    counts: dict[str, int] = {}
    for record in records:
        counts[record.source_file] = counts.get(record.source_file, 0) + 1
    for source_file, count in sorted(counts.items()):
        ws_summary.append([source_file, count])

    ws_summary.append([])
    ws_summary.append(["Part Validation", ""])
    ws_summary.append(["Source File", "Part Number", "Part Name", "Order Qty", "Serial Count", "Qty Check", "Block Source"])
    for row in part_summary_rows(records):
        ws_summary.append(row)

    if warnings:
        ws_summary.append([])
        ws_summary.append(["Warnings", ""])
        for warning in warnings:
            ws_summary.append([warning, ""])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    mismatch_fill = PatternFill("solid", fgColor="F4CCCC")
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row in sheet.iter_rows(min_row=2):
            if any(cell.value == "MISMATCH" for cell in row):
                for cell in row:
                    cell.fill = mismatch_fill
        autosize_sheet(sheet, get_column_letter)

    wb.save(output_xlsx)

    output_csv = output_xlsx.with_suffix(".csv")
    with output_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(output_headers)
        for record in records:
            writer.writerow(
                [
                    record.part_number,
                    record.part_name,
                    record.serial_number,
                    "" if record.order_qty is None else record.order_qty,
                    "" if record.serial_count is None else record.serial_count,
                    record.qty_check,
                ]
            )

    return output_xlsx, output_csv


def run_extraction(
    input_paths: list[Path],
    output_xlsx: Path,
    log: Callable[[str], None] = log_noop,
    options: ExtractionOptions | None = None,
    ocr_engine: OcrEngine | None = None,
) -> tuple[Path, Path, int, list[str]]:
    pdf_paths = collect_pdf_paths(input_paths)
    if not pdf_paths:
        raise ValueError("No PDF files selected.")

    all_records: list[SerialRecord] = []
    warnings: list[str] = []
    reviews: list[ReviewCandidate] = []
    diagnostics: dict[str, object] = {
        "ocr_page_count": 0,
    }
    ocr_status = OcrSupportManager().status()
    diagnostics["ocr_installed"] = ocr_status.installed
    diagnostics["ocr_version"] = ocr_status.version
    for pdf_path in pdf_paths:
        result = extract_pdf(
            pdf_path,
            log=log,
            options=options,
            ocr_engine=ocr_engine,
        )
        all_records.extend(result.records)
        warnings.extend(result.warnings)
        reviews.extend(result.review_candidates)
        diagnostics["ocr_page_count"] = int(diagnostics["ocr_page_count"]) + int(
            result.diagnostics.get("ocr_page_count", 0)
        )

    apply_quantity_validation(all_records, warnings)
    if reviews:
        warnings.append(f"{len(reviews)} OCR candidate(s) require manual review.")
    output_xlsx, output_csv = write_outputs(
        all_records,
        warnings,
        output_xlsx,
        review_candidates=reviews,
        diagnostics=diagnostics,
    )
    return output_xlsx, output_csv, len(all_records), warnings


def default_output_path(base_dir: Path | None = None) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    root = base_dir or Path.cwd()
    return root / f"serial_numbers_{stamp}.xlsx"


def execute_extraction_job(
    input_paths: list[Path],
    output_dir: Path,
    options: ExtractionOptions,
    log: Callable[[str], None] = log_noop,
) -> tuple[Path, Path, int, list[str]]:
    return run_extraction(
        input_paths,
        default_output_path(output_dir),
        log=log,
        options=options,
    )


def launch_gui() -> None:  # pragma: no cover - Tk drawing is manually verified; orchestration is tested separately.
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title(APP_NAME)
    root.geometry("840x560")
    root.minsize(760, 480)

    selected_files: list[Path] = []
    output_dir = tk.StringVar(value=str(Path.cwd() / "outputs"))
    ocr_mode = tk.StringVar(value="auto")
    ui_actions: queue.SimpleQueue[Callable[[], None]] = queue.SimpleQueue()

    def post_ui(action: Callable[[], None]) -> None:
        ui_actions.put(action)

    def drain_ui_actions() -> None:
        while True:
            try:
                ui_actions.get_nowait()()
            except queue.Empty:
                break
        root.after(50, drain_ui_actions)

    def refresh_list() -> None:
        listbox.delete(0, tk.END)
        for path in selected_files:
            listbox.insert(tk.END, str(path))

    def add_files() -> None:
        filenames = filedialog.askopenfilenames(
            title="Select PDF files",
            filetypes=[("PDF files", "*.pdf *.PDF"), ("All files", "*.*")],
        )
        for filename in filenames:
            path = Path(filename)
            if path not in selected_files:
                selected_files.append(path)
        refresh_list()

    def add_folder() -> None:
        folder = filedialog.askdirectory(title="Select folder containing PDF files")
        if folder:
            path = Path(folder)
            if path not in selected_files:
                selected_files.append(path)
            refresh_list()

    def clear_files() -> None:
        selected_files.clear()
        refresh_list()

    def remove_selected() -> None:
        indexes = list(listbox.curselection())
        for index in reversed(indexes):
            selected_files.pop(index)
        refresh_list()

    def choose_output_dir() -> None:
        folder = filedialog.askdirectory(title="Select output folder")
        if folder:
            output_dir.set(folder)

    def log(message: str) -> None:
        def append() -> None:
            log_box.configure(state="normal")
            log_box.insert(tk.END, message + "\n")
            log_box.see(tk.END)
            log_box.configure(state="disabled")

        post_ui(append)

    def install_ocr_support() -> None:
        choice = messagebox.askyesnocancel(
            APP_NAME,
            "Install local PaddleOCR support?\n\n"
            "Yes: download the signed support package.\n"
            "No: select an offline OCR support zip.\n"
            "Cancel: do not install.",
        )
        if choice is None:
            return
        package_path: Path | None = None
        if choice is False:
            selected = filedialog.askopenfilename(
                title="Select OCR support package",
                filetypes=[("ZIP package", "*.zip"), ("All files", "*.*")],
            )
            if not selected:
                return
            package_path = Path(selected)

        ocr_install_button.configure(state="disabled")

        def worker() -> None:
            try:
                manager = OcrSupportManager()
                if package_path is None:
                    installed = manager.install_from_manifest_url()
                else:
                    installed = manager.install_from_package(package_path)
                post_ui(
                    lambda: messagebox.showinfo(
                        APP_NAME,
                        f"OCR support {installed.version} installed.\n\n{installed.worker_path}",
                    )
                )
            except Exception as exc:
                post_ui(
                    lambda exc=exc: messagebox.showerror(
                        APP_NAME,
                        f"OCR installation failed:\n{exc}",
                    )
                )
            finally:
                post_ui(lambda: ocr_install_button.configure(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def run_worker() -> None:
        try:
            out_dir = Path(output_dir.get())
            log("Starting extraction...")
            options = ExtractionOptions(ocr_mode=ocr_mode.get())
            xlsx, csv_path, count, warnings = execute_extraction_job(
                selected_files,
                out_dir,
                options=options,
                log=log,
            )
            log("")
            log(f"Done. Extracted {count} serial numbers.")
            log(f"Excel: {xlsx}")
            log(f"CSV:   {csv_path}")
            if warnings:
                log("")
                log("Warnings:")
                for warning in warnings:
                    log(f"- {warning}")
            post_ui(
                lambda: messagebox.showinfo(
                    APP_NAME,
                    f"Done. Extracted {count} serial numbers.\n\n{xlsx}",
                )
            )
        except OcrSupportRequired as exc:
            log(f"OCR support required: {exc}")
            post_ui(install_ocr_support)
        except Exception as exc:
            log(f"ERROR: {exc}")
            post_ui(lambda exc=exc: messagebox.showerror(APP_NAME, str(exc)))
        finally:
            post_ui(lambda: extract_button.configure(state="normal"))

    def extract() -> None:
        if not selected_files:
            messagebox.showwarning(APP_NAME, "Please add PDF files or a folder first.")
            return
        extract_button.configure(state="disabled")
        threading.Thread(target=run_worker, daemon=True).start()

    main = ttk.Frame(root, padding=12)
    main.pack(fill=tk.BOTH, expand=True)

    title = ttk.Label(main, text=APP_NAME, font=("Segoe UI", 16, "bold"))
    title.pack(anchor="w")
    subtitle = ttk.Label(
        main,
        text="Extract vendor PDF part numbers, part names, serial numbers, and quantity checks into Excel.",
    )
    subtitle.pack(anchor="w", pady=(0, 10))

    button_row = ttk.Frame(main)
    button_row.pack(fill=tk.X)
    ttk.Button(button_row, text="Add PDFs", command=add_files).pack(side=tk.LEFT, padx=(0, 6))
    ttk.Button(button_row, text="Add Folder", command=add_folder).pack(side=tk.LEFT, padx=(0, 6))
    ttk.Button(button_row, text="Remove Selected", command=remove_selected).pack(side=tk.LEFT, padx=(0, 6))
    ttk.Button(button_row, text="Clear", command=clear_files).pack(side=tk.LEFT, padx=(0, 6))

    listbox_frame = ttk.Frame(main)
    listbox_frame.pack(fill=tk.BOTH, expand=True, pady=10)
    listbox = tk.Listbox(listbox_frame, selectmode=tk.EXTENDED)
    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=listbox.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    listbox.configure(yscrollcommand=scrollbar.set)

    output_row = ttk.Frame(main)
    output_row.pack(fill=tk.X, pady=(0, 10))
    ttk.Label(output_row, text="Output folder:").pack(side=tk.LEFT)
    ttk.Entry(output_row, textvariable=output_dir).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8)
    ttk.Button(output_row, text="Browse", command=choose_output_dir).pack(side=tk.LEFT)

    ocr_row = ttk.Frame(main)
    ocr_row.pack(fill=tk.X, pady=(0, 10))
    ttk.Label(ocr_row, text="OCR:").pack(side=tk.LEFT)
    ttk.Combobox(
        ocr_row,
        textvariable=ocr_mode,
        values=("auto", "off", "force"),
        state="readonly",
        width=10,
    ).pack(side=tk.LEFT, padx=8)
    ocr_install_button = ttk.Button(
        ocr_row,
        text="Install OCR Support",
        command=install_ocr_support,
    )
    ocr_install_button.pack(side=tk.LEFT)

    extract_button = ttk.Button(main, text="Extract to Excel", command=extract)
    extract_button.pack(anchor="w", pady=(0, 10))

    log_box = tk.Text(main, height=9, state="disabled")
    log_box.pack(fill=tk.BOTH, expand=False)

    root.after(50, drain_ui_actions)
    root.mainloop()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Extract vendor PDF serial numbers into Excel.")
    parser.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    parser.add_argument("inputs", nargs="*", help="PDF files or folders.")
    parser.add_argument("-o", "--output", help="Output .xlsx path.")
    parser.add_argument("--no-gui", action="store_true", help="Run in command line mode.")
    parser.add_argument(
        "--ocr",
        choices=("auto", "off", "force"),
        default="auto",
        help="OCR mode. Auto only processes pages without usable native text.",
    )
    parser.add_argument(
        "--ocr-min-confidence",
        type=float,
        default=0.90,
        help="Minimum PaddleOCR confidence for main output; lower candidates go to Review.",
    )
    parser.add_argument(
        "--install-ocr",
        action="store_true",
        help="Install optional PaddleOCR support and exit.",
    )
    parser.add_argument(
        "--ocr-package",
        help="Offline OCR support zip for --install-ocr.",
    )
    parser.add_argument(
        "--ocr-manifest-url",
        default=DEFAULT_OCR_MANIFEST_URL,
        help=argparse.SUPPRESS,
    )
    args = parser.parse_args(argv)

    if args.install_ocr or args.ocr_package:
        try:
            manager = OcrSupportManager()
            if args.ocr_package:
                installed = manager.install_from_package(Path(args.ocr_package))
            else:
                installed = manager.install_from_manifest_url(args.ocr_manifest_url)
            cli_log(f"OCR support {installed.version} installed.")
            cli_log(f"Worker: {installed.worker_path}")
            return 0
        except OcrInstallError as exc:
            cli_log(f"ERROR: {exc}")
            return 2

    if not args.no_gui and not args.inputs:
        launch_gui()
        return 0

    input_paths = [Path(value) for value in args.inputs]
    output_path = Path(args.output) if args.output else default_output_path(Path.cwd() / "outputs")
    try:
        options = ExtractionOptions(
            ocr_mode=args.ocr,
            ocr_min_confidence=args.ocr_min_confidence,
        )
        xlsx, csv_path, count, warnings = run_extraction(
            input_paths,
            output_path,
            log=cli_log,
            options=options,
        )
    except OcrSupportRequired as exc:
        cli_log(f"OCR support required: {exc}")
        cli_log("Run again with --install-ocr, or use --ocr off to skip scanned pages.")
        return 3
    except Exception as exc:
        error_log = write_error_log(output_path, exc)
        cli_log(f"ERROR: {exc}")
        cli_log(f"Error log: {error_log}")
        return 1
    cli_log(f"Done. Extracted {count} serial numbers.")
    cli_log(f"Excel: {xlsx}")
    cli_log(f"CSV:   {csv_path}")
    if warnings:
        cli_log("Warnings:")
        for warning in warnings:
            cli_log(f"- {warning}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
