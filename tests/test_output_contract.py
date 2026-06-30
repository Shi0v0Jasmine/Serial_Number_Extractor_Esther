from __future__ import annotations

import csv

from openpyxl import load_workbook
import pytest
from reportlab.pdfgen import canvas

from serial_extractor.models import ExtractionOptions, ReviewCandidate
from serial_number_extractor import SerialRecord, run_extraction, write_outputs


def test_xlsx_and_csv_contract(tmp_path) -> None:
    records = [
        SerialRecord(
            source_file="sample.pdf",
            page=1,
            part_number="PART-001",
            part_name="Example Part",
            order_qty=1,
            serial_number="FA70000000001",
            serial_count=1,
            qty_check="OK",
            item_hint="PART-001",
            method="known_pattern",
            confidence="high",
            block_key="block-1",
            block_source="test",
        )
    ]
    reviews = [
        ReviewCandidate(
            source_file="scan.pdf",
            page=2,
            candidate_type="serial",
            raw_text="FA7O00000001",
            normalized_value="FA7O00000001",
            confidence=0.72,
            reason="ocr_confidence_below_threshold",
            backend="paddleocr",
            strategy="ocr_candidate_filter",
        )
    ]
    xlsx_path, csv_path = write_outputs(
        records,
        [],
        tmp_path / "result.xlsx",
        review_candidates=reviews,
        diagnostics={"ocr_page_count": 1},
    )

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    assert workbook.sheetnames == ["SAP_Copy", "Details", "Review", "Summary"]
    rows = list(workbook["SAP_Copy"].iter_rows(values_only=True))
    assert rows[0] == (
        "Part Number",
        "Part Name",
        "Serial Number",
        "Order Qty",
        "Serial Count",
        "Qty Check",
    )
    assert rows[1] == ("PART-001", "Example Part", "FA70000000001", 1, 1, "OK")
    detail_headers = [cell.value for cell in workbook["Details"][1]]
    assert "Backend" in detail_headers
    assert "Strategy" in detail_headers
    assert "OCR Confidence" in detail_headers
    review_rows = list(workbook["Review"].iter_rows(values_only=True))
    assert review_rows[1][0:7] == (
        "scan.pdf",
        2,
        "serial",
        "FA7O00000001",
        "FA7O00000001",
        0.72,
        "ocr_confidence_below_threshold",
    )
    summary_values = {
        row[0]: row[1]
        for row in workbook["Summary"].iter_rows(min_row=1, max_col=2, values_only=True)
        if row[0]
    }
    assert summary_values["Review Candidates"] == 1
    assert summary_values["OCR Records"] == 0
    assert summary_values["OCR Pages"] == 1
    assert summary_values["OCR Support Installed"] is False
    workbook.close()

    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        csv_rows = list(csv.reader(handle))
    assert csv_rows[0] == list(rows[0])
    assert csv_rows[1] == ["PART-001", "Example Part", "FA70000000001", "1", "1", "OK"]


def test_run_extraction_synthetic_pdf_writes_xlsx_and_csv(tmp_path) -> None:
    pdf_path = tmp_path / "synthetic_adtran.pdf"
    pdf = canvas.Canvas(str(pdf_path), pagesize=(612, 792))
    y = 750
    for line in [
        "500 BC00000647",
        "F7/9TCE-PCN-10GU+10G&1P-L",
        "This Position Line Contains:",
        "Qty Material Mat. Desc. / Customer Mat. Desc.",
        "2 1063707680-11 F7/9TCE-PCN-10GU+10G",
        "S/N:",
        "FA70000000001",
        "FA70000000002",
    ]:
        pdf.drawString(72, y, line)
        y -= 18
    pdf.save()

    xlsx_path, csv_path, count, warnings = run_extraction(
        [tmp_path],
        tmp_path / "result.xlsx",
        options=ExtractionOptions(ocr_mode="off"),
    )

    assert count == 2
    assert warnings == []
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = list(workbook["SAP_Copy"].iter_rows(values_only=True))
    workbook.close()
    assert rows == [
        ("Part Number", "Part Name", "Serial Number", "Order Qty", "Serial Count", "Qty Check"),
        ("BC00000647", "F7/9TCE-PCN-10GU+10G&1P-L", "FA70000000001", 2, 2, "OK"),
        ("BC00000647", "F7/9TCE-PCN-10GU+10G&1P-L", "FA70000000002", 2, 2, "OK"),
    ]

    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        assert list(csv.reader(handle)) == [
            ["Part Number", "Part Name", "Serial Number", "Order Qty", "Serial Count", "Qty Check"],
            ["BC00000647", "F7/9TCE-PCN-10GU+10G&1P-L", "FA70000000001", "2", "2", "OK"],
            ["BC00000647", "F7/9TCE-PCN-10GU+10G&1P-L", "FA70000000002", "2", "2", "OK"],
        ]


def test_run_extraction_rejects_inputs_without_pdfs(tmp_path) -> None:
    note = tmp_path / "not-a-pdf.txt"
    note.write_text("not a pdf", encoding="utf-8")

    with pytest.raises(ValueError, match="No PDF files selected"):
        run_extraction([note], tmp_path / "result.xlsx")
