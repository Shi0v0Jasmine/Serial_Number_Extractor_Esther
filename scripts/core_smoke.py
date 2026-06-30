from __future__ import annotations

import csv
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook
from reportlab.pdfgen import canvas

from serial_extractor.app import apply_quantity_validation, extract_document_records, run_extraction
from serial_extractor.layout import extract_layout_groups
from serial_extractor.models import DocumentPage, ExtractionOptions, TextSpan
from serial_extractor.ocr_worker import parse_paddle_result


def _span(text: str, x: float, y: float) -> TextSpan:
    return TextSpan(text, 1, (x, y, x + max(20, len(text) * 6), y + 10))


def _write_synthetic_pdf(path: Path) -> None:
    pdf = canvas.Canvas(str(path), pagesize=(612, 792))
    y = 750
    for line in [
        "500 BC00000647",
        "F7/9TCE-PCN-10GU+10G&1P-L",
        "This Position Line Contains:",
        "Qty Material Mat. Desc. / Customer Mat. Desc.",
        "2 1063707680-11 F7/9TCE-PCN-10GU+10G",
        "S/N:",
        "FA70000000001",
        "FA70000000002",
    ]:
        pdf.drawString(72, y, line)
        y -= 18
    pdf.save()


def main() -> int:
    adtran_serials = [f"FA7000000{value:04d}" for value in range(1, 14)]
    adtran_lines = [
        "500 BC00000647",
        "F7/9TCE-PCN-10GU+10G&1P-L",
        "This Position Line Contains:",
        "Qty Material Description",
        "13 1063707680-11 F7/9TCE-PCN-10GU+10G",
        "S/N:",
        *adtran_serials,
    ]
    records = extract_document_records(
        "adtran.pdf",
        adtran_lines,
        [1] * len(adtran_lines),
        set(),
    )
    warnings: list[str] = []
    apply_quantity_validation(records, warnings)
    assert len(records) == 13
    assert {record.part_number for record in records} == {"BC00000647"}
    assert {record.qty_check for record in records} == {"OK"}

    alias_lines = [
        "1 1042004437-01",
        "PART-A",
        "S/N:",
        "FA70000000001",
        "USI Code:",
        "LBADVA70000000001",
        "1 1 0 10.00 10.00",
        "2 1042004438-01",
        "PART-B",
        "USI Code:",
        "LBADVA70000000002",
    ]
    aliases = extract_document_records(
        "alias.pdf",
        alias_lines,
        [1] * len(alias_lines),
        set(),
    )
    assert [record.serial_number for record in aliases] == [
        "FA70000000001",
        "LBADVA70000000002",
    ]

    spans = [
        _span("S/N", 10, 100),
        _span("Part Number", 180, 100),
        _span("Qty", 380, 100),
        _span("Description", 470, 100),
        _span("ZX9A001", 10, 130),
        _span("PN-001", 180, 130),
        _span("1", 380, 130),
        _span("Generic Module", 470, 130),
    ]
    groups = extract_layout_groups([DocumentPage(1, "", [], spans)])
    assert [
        (group.part_number, group.part_name, group.order_qty, group.serials)
        for group in groups
    ] == [("PN-001", "Generic Module", 1, ["ZX9A001"])]

    paddle = parse_paddle_result(
        {
            "rec_texts": ["FA7O2612O1473"],
            "rec_scores": [0.73],
            "rec_boxes": [[1, 2, 101, 22]],
        },
        1,
    )
    assert paddle[0]["text"] == "FA7O2612O1473"
    assert paddle[0]["confidence"] == 0.73

    with tempfile.TemporaryDirectory() as tmp:
        workdir = Path(tmp)
        pdf_path = workdir / "synthetic_adtran.pdf"
        output_xlsx = workdir / "serial_numbers.xlsx"
        _write_synthetic_pdf(pdf_path)
        written_xlsx, written_csv, count, export_warnings = run_extraction(
            [pdf_path],
            output_xlsx,
            options=ExtractionOptions(ocr_mode="off"),
        )
        assert count == 2
        assert not export_warnings
        workbook = load_workbook(written_xlsx, read_only=True, data_only=True)
        assert workbook.sheetnames == ["SAP_Copy", "Details", "Review", "Summary"]
        rows = list(workbook["SAP_Copy"].iter_rows(values_only=True))
        workbook.close()
        assert rows == [
            ("Part Number", "Part Name", "Serial Number", "Order Qty", "Serial Count", "Qty Check"),
            ("BC00000647", "F7/9TCE-PCN-10GU+10G&1P-L", "FA70000000001", 2, 2, "OK"),
            ("BC00000647", "F7/9TCE-PCN-10GU+10G&1P-L", "FA70000000002", 2, 2, "OK"),
        ]
        with written_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            csv_rows = list(csv.reader(handle))
        assert csv_rows == [
            ["Part Number", "Part Name", "Serial Number", "Order Qty", "Serial Count", "Qty Check"],
            ["BC00000647", "F7/9TCE-PCN-10GU+10G&1P-L", "FA70000000001", "2", "2", "OK"],
            ["BC00000647", "F7/9TCE-PCN-10GU+10G&1P-L", "FA70000000002", "2", "2", "OK"],
        ]

    print("Core smoke passed: Adtran, alias scope, generic layout, Paddle adapter, XLSX/CSV export")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
